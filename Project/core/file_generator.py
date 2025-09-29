"""
파일 생성 모듈 (Excel/CSV)
검색 결과를 기반으로 다운로드 파일 생성
"""

# pandas removed to reduce serverless function size
import json
from typing import List, Dict, Any, Optional
from pathlib import Path
import asyncio
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# dataframe_to_rows removed with pandas
from datetime import datetime
import io
import csv

class FileGenerator:
    """파일 생성기"""
    
    def __init__(self):
        self.progress_callbacks = {}
    
    def set_progress_callback(self, temp_id: str, callback):
        """진행률 콜백 설정"""
        self.progress_callbacks[temp_id] = callback
    
    def _update_progress(self, temp_id: str, progress: int, message: str = ""):
        """진행률 업데이트"""
        callback = self.progress_callbacks.get(temp_id)
        if callback:
            callback(progress, message)
    
    async def generate_excel(self, 
                           data: List[Dict[str, Any]], 
                           temp_id: str,
                           file_path: Path,
                           metadata: Optional[Dict] = None) -> bool:
        """Excel 파일 생성"""
        try:
            self._update_progress(temp_id, 0, "Excel 파일 생성 시작")
            
            # 데이터 준비
            if not data:
                raise ValueError("생성할 데이터가 없습니다")
            
            # 배열 필드를 문자열로 변환 (한글 필드명 적용)
            korean_field_mapping = metadata.get("korean_field_mapping", {}) if metadata else {}
            processed_data = self._process_data_for_excel(data, korean_field_mapping)
            if processed_data:
                headers = list(processed_data[0].keys())
            else:
                headers = []
            self._update_progress(temp_id, 20, "데이터 처리 중")
            
            # Excel 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "검색 결과"
            
            # 헤더 스타일 정의
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 테두리 스타일
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            self._update_progress(temp_id, 40, "Excel 스타일 적용 중")
            
            # 메타데이터 시트 추가 (요약 정보)
            if metadata:
                summary_ws = wb.create_sheet("검색 요약", 0)
                self._add_summary_sheet(summary_ws, metadata, processed_data)
            
            self._update_progress(temp_id, 50, "데이터 입력 중")
            
            # 데이터 입력 (pandas 없이 직접 처리)
            if headers:
                # 헤더 입력
                for c_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=c_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # 데이터 입력
                for r_idx, row_data in enumerate(processed_data, 2):  # 2부터 시작 (헤더 다음)
                    for c_idx, header in enumerate(headers, 1):
                        value = row_data.get(header, "")
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = thin_border
                        
                        # 날짜 형식 처리
                        if isinstance(value, str) and self._is_date_string(value):
                            cell.number_format = 'YYYY-MM-DD'
            
            self._update_progress(temp_id, 70, "열 너비 자동 조정 중")
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # 최대 50자
                ws.column_dimensions[column_letter].width = adjusted_width
            
            self._update_progress(temp_id, 85, "필터 설정 중")
            
            # 자동 필터 추가
            if len(processed_data) > 0 and headers:
                ws.auto_filter.ref = f"A1:{ws.cell(row=len(processed_data)+1, column=len(headers)).coordinate}"
            
            self._update_progress(temp_id, 95, "파일 저장 중")
            
            # 파일 저장
            print(f"Excel 파일 저장 시작 {temp_id}: 경로={file_path}, 부모디렉토리={file_path.parent}")
            
            # 디렉토리 존재 확인 및 생성
            file_path.parent.mkdir(parents=True, exist_ok=True)
            print(f"디렉토리 생성 완료 {temp_id}: {file_path.parent}")
            
            wb.save(file_path)
            
            # 파일 생성 확인
            if file_path.exists():
                file_size = file_path.stat().st_size
                print(f"Excel 파일 저장 성공 {temp_id}: 크기={file_size} bytes")
                self._update_progress(temp_id, 100, "Excel 파일 생성 완료")
                return True
            else:
                print(f"Excel 파일 저장 실패 {temp_id}: wb.save() 완료했지만 파일이 존재하지 않음")
                self._update_progress(temp_id, -1, "Excel 저장 후 파일 없음")
                return False
            
        except Exception as e:
            print(f"Excel 생성 중 예외 발생 {temp_id}: {e}")
            self._update_progress(temp_id, -1, f"Excel 생성 실패: {str(e)}")
            return False
    
    def _add_summary_sheet(self, ws, metadata: Dict, data: List[Dict]):
        """요약 시트 추가"""
        try:
            # 제목
            ws['A1'] = "DataPage 검색 결과 요약"
            ws['A1'].font = Font(size=16, bold=True)
            
            row = 3
            
            # 기본 정보
            ws[f'A{row}'] = "생성 일시:"
            ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            row += 1
            
            ws[f'A{row}'] = "총 데이터 수:"
            ws[f'B{row}'] = len(data)
            row += 1
            
            # 검색 조건
            if metadata.get('search_conditions'):
                row += 1
                ws[f'A{row}'] = "검색 조건:"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                conditions = metadata['search_conditions']
                for key, value in conditions.items():
                    if value:
                        ws[f'A{row}'] = f"  {key}:"
                        ws[f'B{row}'] = str(value)
                        row += 1
            
            # 카테고리 분포 (데이터에 카테고리가 있는 경우)
            if data and any('category' in item for item in data):
                row += 1
                ws[f'A{row}'] = "카테고리별 분포:"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                # 카테고리 집계
                category_counts = {}
                for item in data:
                    if 'category' in item and item['category']:
                        category = item['category']
                        category_counts[category] = category_counts.get(category, 0) + 1
                
                for category, count in category_counts.items():
                    ws[f'A{row}'] = f"  {category}:"
                    ws[f'B{row}'] = count
                    row += 1
            
        except Exception as e:
            print(f"요약 시트 생성 실패: {e}")
    
    async def generate_csv(self, 
                          data: List[Dict[str, Any]], 
                          temp_id: str,
                          file_path: Path,
                          metadata: Optional[Dict] = None) -> bool:
        """CSV 파일 생성"""
        try:
            self._update_progress(temp_id, 0, "CSV 파일 생성 시작")
            
            if not data:
                raise ValueError("생성할 데이터가 없습니다")
            
            self._update_progress(temp_id, 30, "데이터 처리 중")
            
            # 배열 필드를 문자열로 변환 (한글 필드명 적용)
            korean_field_mapping = metadata.get("korean_field_mapping", {}) if metadata else {}
            processed_data = self._process_data_for_excel(data, korean_field_mapping)
            
            self._update_progress(temp_id, 70, "CSV 파일 저장 중")
            
            # CSV 저장 (UTF-8 BOM 추가로 한글 호환성 확보) - pandas 없이 구현
            import csv
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                if processed_data:
                    fieldnames = list(processed_data[0].keys())
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(processed_data)
            
            self._update_progress(temp_id, 100, "CSV 파일 생성 완료")
            return True
            
        except Exception as e:
            self._update_progress(temp_id, -1, f"CSV 생성 실패: {str(e)}")
            return False
    
    async def generate_json(self, 
                           data: List[Dict[str, Any]], 
                           temp_id: str,
                           file_path: Path,
                           metadata: Optional[Dict] = None) -> bool:
        """JSON 파일 생성"""
        try:
            self._update_progress(temp_id, 0, "JSON 파일 생성 시작")
            
            if not data:
                raise ValueError("생성할 데이터가 없습니다")
            
            self._update_progress(temp_id, 30, "데이터 처리 중")
            
            # 메타데이터와 함께 JSON 구성
            output_data = {
                "metadata": {
                    "generated_at": datetime.now().isoformat(),
                    "total_count": len(data),
                    "generator": "DataPage v1.0"
                },
                "search_conditions": metadata.get('search_conditions', {}) if metadata else {},
                "data": data
            }
            
            # 추가 메타데이터 병합
            if metadata:
                output_data["metadata"].update(metadata)
            
            self._update_progress(temp_id, 70, "JSON 파일 저장 중")
            
            # JSON 저장
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, ensure_ascii=False, indent=2)
            
            self._update_progress(temp_id, 100, "JSON 파일 생성 완료")
            return True
            
        except Exception as e:
            self._update_progress(temp_id, -1, f"JSON 생성 실패: {str(e)}")
            return False
    
    def _process_data_for_excel(self, data: List[Dict[str, Any]], korean_field_mapping: Dict[str, str] = None) -> List[Dict[str, Any]]:
        """Excel 출력을 위한 데이터 전처리 (한글 필드명 적용)"""
        processed_data = []
        
        for item in data:
            processed_item = {}
            for key, value in item.items():
                # 한글 필드명이 있으면 사용, 없으면 원본 사용
                display_key = korean_field_mapping.get(key, key) if korean_field_mapping else key
                
                if isinstance(value, list):
                    # 배열을 쉼표로 구분된 문자열로 변환
                    processed_item[display_key] = ", ".join(str(v) for v in value)
                elif isinstance(value, dict):
                    # 딕셔너리를 JSON 문자열로 변환
                    processed_item[display_key] = json.dumps(value, ensure_ascii=False)
                else:
                    processed_item[display_key] = value
            processed_data.append(processed_item)
        
        return processed_data
    
    def _is_date_string(self, value: str) -> bool:
        """날짜 문자열 여부 확인"""
        try:
            if isinstance(value, str) and len(value) >= 10:
                # YYYY-MM-DD 형식 확인
                datetime.strptime(value[:10], '%Y-%m-%d')
                return True
        except ValueError:
            pass
        return False
    
    def estimate_file_size(self, data: List[Dict[str, Any]], file_type: str, total_records: Optional[int] = None) -> int:
        """파일 크기 추정 (바이트)"""
        if not data and not total_records:
            return 0
        
        row_count = total_records if total_records is not None else len(data)
        if row_count == 0:
            return 0

        # 샘플 데이터로 크기 추정
        sample_size = min(len(data), 100)
        sample_data = data[:sample_size]

        try:
            if file_type.lower() == 'json':
                sample_json = json.dumps(sample_data, ensure_ascii=False)
                estimated_size = len(sample_json.encode('utf-8')) * (row_count / sample_size)
                return int(estimated_size * 1.2)  # 메타데이터 등 여유분
            
            elif file_type.lower() == 'csv':
                output = io.StringIO()
                import csv
                if sample_data:
                    fieldnames = list(sample_data[0].keys())
                    writer = csv.DictWriter(output, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(sample_data)
                    sample_size_bytes = len(output.getvalue().encode('utf-8'))
                    estimated_size = sample_size_bytes * (row_count / sample_size)
                    return int(estimated_size * 1.1)  # 여유분
                return 0
            
            elif file_type.lower() in ['xlsx', 'excel']:
                # Excel은 압축되므로 CSV보다 작지만, 스타일링으로 인해 더 클 수 있음
                output = io.StringIO()
                import csv
                if sample_data:
                    fieldnames = list(sample_data[0].keys())
                    writer = csv.DictWriter(output, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(sample_data)
                    csv_size = len(output.getvalue().encode('utf-8'))
                    estimated_size = csv_size * (row_count / sample_size) * 1.5  # Excel 오버헤드
                    return int(estimated_size)
                return 0
            
        except Exception as e:
            print(f"파일 크기 추정 실패: {e}")
            # 기본 추정치 (레코드당 평균 500바이트)
            return len(data) * 500
        
        return len(data) * 500  # 기본값

# 전역 인스턴스
file_generator = FileGenerator()
