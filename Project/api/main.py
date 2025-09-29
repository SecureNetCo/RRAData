"""
DataPage FastAPI 서버 - Vercel 서버리스 환경 최적화
"""

# Vercel 서버리스 환경에서 모듈 경로 설정
import sys
import os
from pathlib import Path
import urllib.request
import shutil
import threading

# 현재 파일의 프로젝트 루트 경로를 Python path에 추가
current_dir = Path(__file__).parent
project_root = current_dir.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

print(f"Python path에 추가된 경로: {project_root}")
print(f"현재 작업 디렉토리: {os.getcwd()}")
print(f"sys.path: {sys.path[:3]}...")  # 처음 3개만 출력

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any, Tuple, Union
import json
import asyncio
import logging
from datetime import datetime
# pandas removed to reduce serverless function size
import tempfile
from urllib.parse import urlparse, quote
from openpyxl import Workbook

# 로깅 설정 - Vercel 환경에 최적화
logging.basicConfig(
    level=logging.INFO,
    format='%(levelname)s:%(name)s:%(message)s',
    force=True  # 기존 설정 덮어쓰기
)

# 특정 모듈의 과도한 로그 레벨 조정
logging.getLogger('urllib3').setLevel(logging.WARNING)
logging.getLogger('requests').setLevel(logging.WARNING)
logging.getLogger('config.display_config').setLevel(logging.WARNING)  # 설정 로딩 로그 줄이기

logger = logging.getLogger(__name__)

# DuckDB 기본 사용 (모든 검색에 parquet + DuckDB 사용)
USE_DUCKDB = True  # 항상 DuckDB 사용

# Blob 파일 사전 다운로드 및 캐시 경로 설정
BLOB_PREFETCH_ROOT = Path("/tmp/datapage_blobs")
PREFETCH_LOCK = threading.Lock()
PREFETCHED_BLOB_FILES: Dict[Tuple[str, Optional[str], str], str] = {}

# 콜드스타트 정보 저장용 전역 변수
_cold_start_info: Optional[Dict[str, Any]] = None

# 사전 다운로드 대상 Blob 환경변수 매핑 (category, result_type, subcategory) → env var
BLOB_ENV_PREFETCH_MAPPING: Dict[Tuple[str, Optional[str], str], str] = {
    # DataA (2025 경량 데이터)
    ("dataA", None, "safetykorea"): "BLOB_URL_DATAA_1_SAFETYKOREA",
    ("dataA", None, "efficiency-rating"): "BLOB_URL_DATAA_3_EFFICIENCY",
    ("dataA", None, "high-efficiency"): "BLOB_URL_DATAA_4_HIGH_EFFICIENCY",
    ("dataA", None, "standby-power"): "BLOB_URL_DATAA_5_STANDBY_POWER",
    ("dataA", None, "approval"): "BLOB_URL_DATAA_6_APPROVAL",
    ("dataA", None, "declaration-details"): "BLOB_URL_DATAA_7_DECLARE",
    ("dataA", None, "kwtc"): "BLOB_URL_DATAA_8_KWTC",
    ("dataA", None, "recall"): "BLOB_URL_DATAA_9_RECALL",
    ("dataA", None, "safetykoreachild"): "BLOB_URL_DATAA_10_SAFETYKOREACHILD",
    ("dataA", None, "rra-cert"): "BLOB_URL_DATAA_11_RRA_CERT",
    ("dataA", None, "rra-self-cert"): "BLOB_URL_DATAA_12_RRA_SELF_CERT",
    ("dataA", None, "safetykoreahome"): "BLOB_URL_DATAA_13_SAFETYKOREAHOME",

    # DataB
    ("dataB", None, "wadiz-makers"): "BLOB_URL_DATAB_2_WADIZ",

    # DataC Success
    ("dataC", "success", "safetykorea"): "BLOB_URL_DATAC_SUCCESS_1_SAFETYKOREA",
    ("dataC", "success", "wadiz-makers"): "BLOB_URL_DATAC_SUCCESS_2_WADIZ",
    ("dataC", "success", "efficiency-rating"): "BLOB_URL_DATAC_SUCCESS_3_EFFICIENCY",
    ("dataC", "success", "high-efficiency"): "BLOB_URL_DATAC_SUCCESS_4_HIGH_EFFICIENCY",
    ("dataC", "success", "standby-power"): "BLOB_URL_DATAC_SUCCESS_5_STANDBY_POWER",
    ("dataC", "success", "approval"): "BLOB_URL_DATAC_SUCCESS_6_APPROVAL",
    ("dataC", "success", "declaration-details"): "BLOB_URL_DATAC_SUCCESS_7_DECLARE",
    ("dataC", "success", "kwtc"): "BLOB_URL_DATAC_SUCCESS_8_KWTC",
    ("dataC", "success", "recall"): "BLOB_URL_DATAC_SUCCESS_9_RECALL",
    ("dataC", "success", "safetykoreachild"): "BLOB_URL_DATAC_SUCCESS_10_SAFETYKOREACHILD",
    ("dataC", "success", "rra-cert"): "BLOB_URL_DATAC_SUCCESS_11_RRA_CERT",
    ("dataC", "success", "rra-self-cert"): "BLOB_URL_DATAC_SUCCESS_12_RRA_SELF_CERT",
    ("dataC", "success", "safetykoreahome"): "BLOB_URL_DATAC_SUCCESS_13_SAFETYKOREAHOME",

    # DataC Failed
    ("dataC", "failed", "safetykorea"): "BLOB_URL_DATAC_FAILED_1_SAFETYKOREA",
    ("dataC", "failed", "wadiz-makers"): "BLOB_URL_DATAC_FAILED_2_WADIZ",
    ("dataC", "failed", "efficiency-rating"): "BLOB_URL_DATAC_FAILED_3_EFFICIENCY",
    ("dataC", "failed", "high-efficiency"): "BLOB_URL_DATAC_FAILED_4_HIGH_EFFICIENCY",
    ("dataC", "failed", "standby-power"): "BLOB_URL_DATAC_FAILED_5_STANDBY_POWER",
    ("dataC", "failed", "approval"): "BLOB_URL_DATAC_FAILED_6_APPROVAL",
    ("dataC", "failed", "declaration-details"): "BLOB_URL_DATAC_FAILED_7_DECLARE",
    ("dataC", "failed", "kwtc"): "BLOB_URL_DATAC_FAILED_8_KWTC",
    ("dataC", "failed", "recall"): "BLOB_URL_DATAC_FAILED_9_RECALL",
    ("dataC", "failed", "safetykoreachild"): "BLOB_URL_DATAC_FAILED_10_SAFETYKOREACHILD",
    ("dataC", "failed", "rra-cert"): "BLOB_URL_DATAC_FAILED_11_RRA_CERT",
    ("dataC", "failed", "rra-self-cert"): "BLOB_URL_DATAC_FAILED_12_RRA_SELF_CERT",
    ("dataC", "failed", "safetykoreahome"): "BLOB_URL_DATAC_FAILED_13_SAFETYKOREAHOME",
}


SUBCATEGORY_ALIAS_MAP: Dict[str, str] = {
    "approval-details": "approval",
    "rra-certification": "rra-cert",
    "rra-self-conformity": "rra-self-cert",
}


def normalize_subcategory(subcategory: Optional[str]) -> Optional[str]:
    if not subcategory:
        return subcategory
    return SUBCATEGORY_ALIAS_MAP.get(subcategory, subcategory)


def _make_prefetch_key(category: str, subcategory: str, result_type: Optional[str] = None) -> Tuple[str, Optional[str], str]:
    normalized_subcategory = normalize_subcategory(subcategory)
    return (category, result_type, normalized_subcategory)


def _inspect_data_source(data_file_path: str) -> Tuple[str, bool, bool, float]:
    """Return path info, whether it's remote, tabular(parquet/duckdb), and size in MB."""

    data_file_str = str(data_file_path)
    is_remote = data_file_str.startswith(('https://', 'http://'))
    is_tabular = data_file_str.lower().endswith(('.parquet', '.duckdb'))

    if is_remote:
        # Blob/R2 URL은 실제 크기를 조회하기 어렵기 때문에 대용량으로 취급
        return data_file_str, True, is_tabular, 100.0

    local_path = Path(data_file_str)
    size_mb = local_path.stat().st_size / (1024 * 1024) if local_path.exists() else 0.0
    return data_file_str, False, is_tabular, size_mb


def _store_prefetched_blob(category: str, subcategory: str, result_type: Optional[str], local_path: str) -> None:
    key = _make_prefetch_key(category, subcategory, result_type)
    with PREFETCH_LOCK:
        PREFETCHED_BLOB_FILES[key] = local_path


def get_prefetched_blob_path(category: str, subcategory: str, result_type: Optional[str] = None) -> Optional[str]:
    key = _make_prefetch_key(category, subcategory, result_type)
    with PREFETCH_LOCK:
        local_path = PREFETCHED_BLOB_FILES.get(key)
    if local_path and os.path.exists(local_path):
        return local_path
    return None


def _derive_blob_filename(url: str, fallback: str) -> str:
    parsed = urlparse(url)
    candidate = Path(parsed.path).name if parsed.path else ""
    if candidate:
        return candidate
    return fallback


def _prefetch_single_blob(category: str, subcategory: str, result_type: Optional[str], url: str) -> Optional[str]:
    try:
        BLOB_PREFETCH_ROOT.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass

    fallback_name = f"{category}_{result_type or 'default'}_{subcategory}.parquet"
    filename = _derive_blob_filename(url, fallback_name)
    dest_path = BLOB_PREFETCH_ROOT / filename

    if dest_path.exists() and dest_path.stat().st_size > 0:
        logger.info(f"Blob 사전 다운로드 재사용: {filename}")
        _store_prefetched_blob(category, subcategory, result_type, str(dest_path))
        return str(dest_path)

    temp_path = dest_path.with_suffix(dest_path.suffix + ".download")

    try:
        logger.info(f"Blob 사전 다운로드 시작: {url} → {dest_path}")
        with urllib.request.urlopen(url) as response, open(temp_path, "wb") as out_file:
            shutil.copyfileobj(response, out_file)
        os.replace(temp_path, dest_path)
        _store_prefetched_blob(category, subcategory, result_type, str(dest_path))
        logger.info(f"Blob 사전 다운로드 완료: {dest_path}")
        return str(dest_path)
    except Exception as download_error:
        logger.warning(f"Blob 사전 다운로드 실패 ({url}): {download_error}")
        if temp_path.exists():
            try:
                temp_path.unlink()
            except Exception:
                pass
        return None


async def prefetch_blob_files() -> None:
    global _cold_start_info
    import time

    start_time = time.time()
    data_mode = os.getenv("DATA_MODE", "full").lower()
    if data_mode != "2025":
        logger.info(f"Blob 사전 다운로드 스킵 (DATA_MODE={data_mode})")
        return

    if not BLOB_ENV_PREFETCH_MAPPING:
        logger.info("Blob 사전 다운로드 대상 매핑 없음")
        return

    prefetch_tasks = []
    task_keys: List[Tuple[str, Optional[str], str]] = []

    for key, env_var in BLOB_ENV_PREFETCH_MAPPING.items():
        url = os.getenv(env_var)
        if not url:
            logger.debug(f"Blob 사전 다운로드 스킵: 환경변수 {env_var} 미설정")
            continue
        category, result_type, subcategory = key
        task = asyncio.to_thread(_prefetch_single_blob, category, subcategory, result_type, url)
        prefetch_tasks.append(task)
        task_keys.append(key)

    if not prefetch_tasks:
        logger.info("Blob 사전 다운로드 수행할 항목 없음")
        return

    results = await asyncio.gather(*prefetch_tasks, return_exceptions=True)
    success_count = 0
    total_files = len(results)

    for key, result in zip(task_keys, results):
        if isinstance(result, Exception):
            logger.warning(f"Blob 사전 다운로드 예외 {key}: {result}")
        elif result:
            logger.info(f"Blob 사전 다운로드 준비 완료 {key}: {result}")
            success_count += 1
        else:
            logger.warning(f"Blob 사전 다운로드 미완료 {key}")

    end_time = time.time()
    duration = end_time - start_time

    # 콜드스타트 정보 설정
    _cold_start_info = {
        "type": "cold_start_complete",
        "message": f"🚀 콜드스타트 완료: {success_count}/{total_files}개 파일, {duration:.2f}초 소요",
        "stats": {
            "success_count": success_count,
            "total_files": total_files,
            "duration_seconds": round(duration, 2),
            "timestamp": datetime.now().isoformat()
        }
    }

    logger.info(_cold_start_info["message"])

# 로컬 모듈 import
from config.search_config import search_config_manager
from config.display_config import display_config_manager, CategoryDisplayConfig, DisplayField, SearchField
from core.large_file_processor import get_processor, stream_search_large_file, SearchContext
from core.duckdb_processor import duckdb_search_large_file


app = FastAPI(title="DataPage API", version="1.0.0")

# CORS 설정
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# **성능 최적화: Startup Warming**
@app.on_event("startup")
async def startup_warming():
    """
    서버 시작시 DuckDB 및 주요 컴포넌트 사전 로드
    - DuckDB 연결 및 httpfs 확장 로드
    - 주요 parquet 파일들 메타데이터 캐싱
    - 첫 번째 사용자의 cold start 방지
    """
    try:
        logger.info("🔥 Startup Warming 시작...")

        # 🧹 모든 모드에서 startup 시 /tmp 정리 수행
        clear_success = clear_tmp_cache()
        if clear_success:
            logger.info("🧹 Startup /tmp 캐시 정리 완료")

        prefetch_config = get_prefetch_config()
        if prefetch_config["enabled"]:
            logger.info("🧭 2025 모드: 스마트 프리페치 활성화, 초기 일괄 다운로드 스킵")
        else:
            await prefetch_blob_files()

        # 주요 카테고리들에 대해 작은 쿼리 실행하여 warming
        # **성능 최적화: Startup Warming - 주요 데이터셋 사전 로딩**
        warming_categories = [
            ("dataA", "safetykorea"),      # 가장 큰 파일
            ("dataA", "safetykoreachild"), # 두 번째로 큰 파일
            ("dataB", "wadiz-makers"),     # 자주 사용되는 파일
        ]

        for category, subcategory in warming_categories:
            try:
                # 각 파일에 대해 최소한의 쿼리 실행 (limit=1)
                warming_request = SearchRequest(
                    keyword="test",
                    search_field="company_name",
                    page=1,
                    limit=1,
                    filters={}
                )
                warming_result = await search_category_data(category, subcategory, warming_request)
                logger.info(f"✅ Warming 완료: {category}_{subcategory}")
            except Exception as e:
                logger.warning(f"⚠️ Warming 실패: {category}_{subcategory} - {e}")

        logger.info("🚀 Startup Warming 완료! 첫 사용자 요청 최적화됨")

    except Exception as e:
        logger.error(f"❌ Startup Warming 전체 실패: {e}")
        # Warming 실패해도 서버는 정상 시작

# Static 파일 경로 설정 - Vercel과 로컬 환경 호환
static_path = project_root / "public" / "static"

# 로컬 환경에서만 static 파일 마운트 (Vercel에서는 vercel.json이 처리)
if os.getenv("VERCEL") is None:  # 로컬 환경에서만
    from fastapi.staticfiles import StaticFiles
    if static_path.exists():
        app.mount("/static", StaticFiles(directory=str(static_path)), name="static")


# 검색 요청 모델
class SearchRequest(BaseModel):
    keyword: Optional[str] = None
    search_field: Optional[str] = "product_name"  # 검색 필드: company_name, model_name, product_name 등 ('all' 제거됨)
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    categories: Optional[List[str]] = None
    filters: Optional[Dict[str, Any]] = None  # 유연한 추가 필터
    # 서버사이드 페이지네이션 파라미터
    page: Optional[int] = 1  # 페이지 번호 (1부터 시작)
    limit: Optional[int] = 20  # 페이지당 항목 수 (기본 20개)
    # offset은 page와 limit으로 계산되므로 제거
    # offset: Optional[int] = 0

# 페이지네이션 정보 모델
class PaginationInfo(BaseModel):
    total_count: int        # 전체 항목 수
    total_pages: int        # 전체 페이지 수
    current_page: int       # 현재 페이지
    items_per_page: int     # 페이지당 항목 수
    has_next: bool          # 다음 페이지 존재 여부
    has_prev: bool          # 이전 페이지 존재 여부

# 검색 응답 모델 (서버사이드 페이지네이션용)
class SearchResponse(BaseModel):
    results: List[Dict[str, Any]]           # 현재 페이지 결과
    pagination: PaginationInfo              # 페이지네이션 정보
    summary: Dict[str, Any]                 # 처리 정보
    available_categories: List[str]         # 사용 가능한 카테고리 (호환성용)

# 다운로드 요청 모델
class DownloadRequest(BaseModel):
    search_conditions: Dict[str, Any]
    file_format: str = "xlsx"  # 엑셀만 지원
    user_session: Optional[str] = None
    filtered_data: Optional[List[Dict[str, Any]]] = None  # 프론트엔드에서 필터링된 데이터

# 검색 기반 전체 다운로드 요청 모델
class SearchDownloadRequest(BaseModel):
    keyword: str
    search_field: Optional[str] = "all"
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    categories: Optional[List[str]] = None
    filters: Optional[Dict[str, Any]] = None
    file_format: str = "xlsx"
    user_session: Optional[str] = None

@app.get("/")
async def root():
    """메인 페이지 - search.html로 리다이렉트"""
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url="/static/search.html", status_code=302)

@app.get("/api/system/status")
async def get_system_status():
    """시스템 상태 및 성능 설정 정보"""
    return {
        "api_version": "1.0.0",
        "duckdb_enabled": USE_DUCKDB,
        "supported_sources": "모든 소스 (Vercel Blob Parquet)",
        "performance_features": {
            "duckdb": "8-11배 성능 향상 (대용량 파일)" if USE_DUCKDB else "비활성화됨",
            "ijson_streaming": "활성화됨",
            "large_file_threshold": "50MB"
        },
        "activation_guide": {
            "duckdb": "환경변수 USE_DUCKDB=true로 설정",
            "github_releases": "1.6GB+ 파일은 GitHub Releases 업로드 권장"
        },
        "prefetch": get_prefetch_config()
    }


@app.get("/search/{category}/{subcategory}")
async def serve_search_page(category: str, subcategory: str):
    """검색 페이지 - search.html 반환 (기존 경로 호환성)"""
    search_path = static_path / "search.html"
    if search_path.exists():
        with open(search_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return HTMLResponse(content=content)
    else:
        raise HTTPException(status_code=404, detail="검색 페이지를 찾을 수 없습니다")

@app.get("/search/dataA/{subcategory}")
async def serve_search_page_data_a(category: str = "dataA", subcategory: str = None):
    """검색 페이지 - dataA 구조"""
    search_path = static_path / "search.html"
    if search_path.exists():
        with open(search_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return HTMLResponse(content=content)
    else:
        raise HTTPException(status_code=404, detail="검색 페이지를 찾을 수 없습니다")

@app.get("/search/dataB/{subcategory}")
async def serve_search_page_data_b(category: str = "dataB", subcategory: str = None):
    """검색 페이지 - dataB 구조"""
    search_path = static_path / "search.html"
    if search_path.exists():
        with open(search_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return HTMLResponse(content=content)
    else:
        raise HTTPException(status_code=404, detail="검색 페이지를 찾을 수 없습니다")

@app.get("/search/dataC/{result_type}/{subcategory}")
async def serve_search_page_data_c(result_type: str, subcategory: str):
    """검색 페이지 - dataC 구조 (success/failed)"""
    search_path = static_path / "search.html"
    if search_path.exists():
        with open(search_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return HTMLResponse(content=content)
    else:
        raise HTTPException(status_code=404, detail="검색 페이지를 찾을 수 없습니다")

@app.get("/admin")
async def serve_admin_page():
    """관리자 페이지 - admin.html 반환"""
    admin_path = static_path / "admin.html"
    if admin_path.exists():
        with open(admin_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return HTMLResponse(content=content)
    else:
        raise HTTPException(status_code=404, detail="관리자 페이지를 찾을 수 없습니다")

# 콜드스타트 정보 조회 엔드포인트
@app.get("/api/cold-start-info")
async def get_cold_start_info():
    """콜드스타트 정보 조회 (개발자 도구 콘솔 표시용)"""
    global _cold_start_info
    if _cold_start_info:
        return _cold_start_info
    else:
        return {
            "type": "cold_start_pending",
            "message": "⏳ 콜드스탄트 진행 중...",
            "stats": None
        }

def clear_tmp_cache():
    """2025 모드용 /tmp 캐시 폴더 정리"""
    try:
        if BLOB_PREFETCH_ROOT.exists():
            import shutil
            shutil.rmtree(BLOB_PREFETCH_ROOT)
            logger.info(f"🧹 /tmp 캐시 폴더 정리 완료: {BLOB_PREFETCH_ROOT}")

        # 메모리 상의 캐시도 정리
        global PREFETCHED_BLOB_FILES
        with PREFETCH_LOCK:
            PREFETCHED_BLOB_FILES.clear()

        return True
    except Exception as e:
        logger.warning(f"⚠️ /tmp 캐시 정리 실패: {e}")
        return False

class SinglePrefetchRequest(BaseModel):
    category: str
    subcategory: str
    result_type: Optional[str] = None

@app.post("/api/prefetch-single")
async def prefetch_single_file(request: SinglePrefetchRequest):
    """2025 모드용 개별 파일 다운로드"""
    import time

    config = get_prefetch_config()
    if not config["enabled"]:
        raise HTTPException(status_code=400, detail="2025 모드에서만 사용 가능합니다")

    start_time = time.time()

    # 1. /tmp 폴더 정리
    clear_success = clear_tmp_cache()

    # 2. 환경변수에서 URL 찾기 (슬러그 정규화)
    normalized_subcategory = normalize_subcategory(request.subcategory)
    key = (request.category, request.result_type, normalized_subcategory)
    env_var = BLOB_ENV_PREFETCH_MAPPING.get(key)

    url = os.getenv(env_var) if env_var else None

    if not url:
        # BLOB 환경변수가 없으면 R2 URL로 fallback 시도
        fallback_url = None
        try:
            if request.category == "dataC" and request.result_type:
                fallback_candidate = get_data_file_path_c(
                    request.category,
                    request.result_type,
                    request.subcategory,
                    prefer_r2=True
                )
            else:
                fallback_candidate = get_data_file_path(
                    request.category,
                    request.subcategory,
                    prefer_r2=True
                )

            if isinstance(fallback_candidate, str) and fallback_candidate.startswith("http"):
                fallback_url = fallback_candidate
                logger.info(
                    f"BLOB URL 없음: {request.category}/{request.result_type}/{request.subcategory} → R2 URL로 프리페치 대체"
                )
        except Exception as fallback_error:
            logger.warning(
                f"프리페치 fallback URL 확인 실패: {fallback_error}"
            )

        if fallback_url:
            url = fallback_url

    if not url:
        raise HTTPException(
            status_code=404,
            detail=f"프리페치 가능한 원격 URL을 찾을 수 없습니다: {request.category}/{request.subcategory}"
        )

    # 3. 개별 파일 다운로드
    try:
        logger.info(f"🎯 개별 파일 다운로드 시작: {request.category}/{request.subcategory}")
        result = await asyncio.to_thread(_prefetch_single_blob, request.category, request.subcategory, request.result_type, url)

        end_time = time.time()
        duration = end_time - start_time

        if result:
            return {
                "success": True,
                "message": f"✅ 파일 다운로드 완료: {request.subcategory}",
                "stats": {
                    "category": request.category,
                    "subcategory": request.subcategory,
                    "result_type": request.result_type,
                    "local_path": result,
                    "duration_seconds": round(duration, 2),
                    "cache_cleared": clear_success,
                    "timestamp": datetime.now().isoformat()
                }
            }
        else:
            raise HTTPException(status_code=500, detail="파일 다운로드 실패")

    except Exception as e:
        logger.error(f"❌ 개별 파일 다운로드 실패: {e}")
        raise HTTPException(status_code=500, detail=f"다운로드 실패: {str(e)}")


@app.get("/api/prefetch/config")
async def get_prefetch_api_config():
    """프리페치 사용 가능 여부(2025 모드) 제공"""
    return get_prefetch_config()

@app.post("/api/search/{category}/{subcategory}")
async def search_category_data(category: str, subcategory: str, request: SearchRequest):
    """
    카테고리별 검색 - DuckDB + Parquet 전용
    """
    try:
        # Parquet 데이터 파일 URL 가져오기
        data_file_path = get_data_file_path(category, subcategory)
        if not data_file_path:
            raise HTTPException(status_code=404, detail=f"데이터 파일 URL을 찾을 수 없습니다: {category}/{subcategory}")
        
        # R2 URL인지 확인하여 적절한 처리 방식 선택
        is_r2_url = data_file_path.startswith('https://') 
        if is_r2_url:
            # R2 URL이면 항상 대용량 파일 처리 (DuckDB) 사용
            file_size_mb = 100.0  # 대용량 처리 로직을 타도록 설정
        else:
            # 로컬 파일이면 실제 크기 확인
            from pathlib import Path
            local_path = Path(data_file_path)
            file_size_mb = local_path.stat().st_size / (1024 * 1024) if local_path.exists() else 0
        
        logger.info(f"DuckDB Parquet 처리 시작: {category}/{subcategory} ({'R2 URL' if is_r2_url else f'{file_size_mb:.1f}MB'})")

        # DuckDB로 Parquet 파일 검색 (페이지네이션)
        effective_subcategory = normalize_subcategory(subcategory)

        search_result = await duckdb_search_large_file(
            file_path=str(data_file_path),
            keyword=request.keyword,
            search_field=request.search_field,
            limit=request.limit,
            page=request.page,
            filters=request.filters,
            category=category,
            subcategory=effective_subcategory
        )
        
        # 오류 발생 시 예외 처리
        if "error" in search_result:
            raise HTTPException(status_code=500, detail=f"검색 처리 실패: {search_result.get('message')}")
        
        # 요약 정보 생성
        summary = {
            "processing_method": "duckdb_pagination",
            "file_size_mb": round(file_size_mb, 2),
            "processing_stats": search_result.get("stats", {}),
            "duckdb_enabled": True,
            "performance_note": "서버사이드 페이지네이션으로 최적화된 처리"
        }

        # 디버그 정보 추가 (search_result에서 가져옴)
        if "debug_info" in search_result and search_result["debug_info"]:
            summary["debug_info"] = search_result["debug_info"]

        # 페이지네이션 정보 생성
        pagination_data = search_result.get("pagination", {})
        pagination_info = PaginationInfo(
            total_count=pagination_data.get("total_count", 0),
            total_pages=pagination_data.get("total_pages", 1),
            current_page=pagination_data.get("current_page", 1),
            items_per_page=pagination_data.get("items_per_page", 20),
            has_next=pagination_data.get("has_next", False),
            has_prev=pagination_data.get("has_prev", False)
        )

        return SearchResponse(
            results=search_result.get("results", []),
            pagination=pagination_info,
            summary=summary,
            available_categories=[]
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"검색 중 오류 발생: {str(e)}")

@app.post("/api/search/dataA/{subcategory}")
async def search_data_a(subcategory: str, request: SearchRequest):
    """
    dataA 카테고리 검색 - 새 구조
    """
    return await search_category_data("dataA", subcategory, request)

@app.post("/api/search/dataB/{subcategory}")
async def search_data_b(subcategory: str, request: SearchRequest):
    """
    dataB 카테고리 검색 - 새 구조
    """
    return await search_category_data("dataB", subcategory, request)

@app.post("/api/search/dataC/{result_type}/{subcategory}")
async def search_data_c(result_type: str, subcategory: str, request: SearchRequest):
    """
    dataC 카테고리 검색 - 새 구조 (success/failed)
    """
    return await search_category_data_c("dataC", result_type, subcategory, request)

async def search_category_data_c(category: str, result_type: str, subcategory: str, request: SearchRequest):
    """
    dataC 카테고리별 검색 - DuckDB + Parquet 전용 (3-parameter structure)
    """
    try:
        # Parquet 데이터 파일 URL 가져오기 (3-parameter structure)
        data_file_path = get_data_file_path_c(category, result_type, subcategory)
        if not data_file_path:
            raise HTTPException(status_code=404, detail=f"데이터 파일 URL을 찾을 수 없습니다: {category}/{result_type}/{subcategory}")
        
        # R2 URL인지 확인하여 적절한 처리 방식 선택
        is_r2_url = data_file_path.startswith('https://') 
        if is_r2_url:
            # R2 URL이면 항상 대용량 파일 처리 (DuckDB) 사용
            file_size_mb = 100.0  # 대용량 처리 로직을 타도록 설정
        else:
            # 로컬 파일이면 실제 크기 확인
            from pathlib import Path
            local_path = Path(data_file_path)
            file_size_mb = local_path.stat().st_size / (1024 * 1024) if local_path.exists() else 0
        
        logger.info(f"DuckDB Parquet 처리 시작: {category}/{result_type}/{subcategory} ({'R2 URL' if is_r2_url else f'{file_size_mb:.1f}MB'})")

        # DuckDB로 Parquet 파일 검색 (페이지네이션)
        effective_subcategory = normalize_subcategory(subcategory)

        search_result = await duckdb_search_large_file(
            file_path=str(data_file_path),
            keyword=request.keyword,
            search_field=request.search_field,
            limit=request.limit,
            page=request.page,
            filters=request.filters,
            category=category,
            subcategory=effective_subcategory,
            result_type=result_type
        )
        
        # 오류 발생 시 예외 처리
        if "error" in search_result:
            raise HTTPException(status_code=500, detail=f"검색 처리 실패: {search_result.get('message')}")
        
        # 요약 정보 생성
        summary = {
            "processing_method": "duckdb_pagination",
            "file_size_mb": round(file_size_mb, 2),
            "processing_stats": search_result.get("stats", {}),
            "duckdb_enabled": True,
            "performance_note": "서버사이드 페이지네이션으로 최적화된 처리"
        }

        # 디버그 정보 추가 (search_result에서 가져옴)
        if "debug_info" in search_result and search_result["debug_info"]:
            summary["debug_info"] = search_result["debug_info"]

        # 페이지네이션 정보 생성
        pagination_data = search_result.get("pagination", {})
        pagination_info = PaginationInfo(
            total_count=pagination_data.get("total_count", 0),
            total_pages=pagination_data.get("total_pages", 1),
            current_page=pagination_data.get("current_page", 1),
            items_per_page=pagination_data.get("items_per_page", 20),
            has_next=pagination_data.get("has_next", False),
            has_prev=pagination_data.get("has_prev", False)
        )

        return SearchResponse(
            results=search_result.get("results", []),
            pagination=pagination_info,
            summary=summary,
            available_categories=[]
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"검색 중 오류 발생: {str(e)}")

@app.post("/api/search")
async def search_data(request: SearchRequest):
    """
    기본 검색 (하위 호환성) - dataA/safetykorea 데이터 사용
    """
    return await search_category_data("dataA", "safetykorea", request)

@app.get("/api/categories")
async def get_categories():
    """
    사용 가능한 모든 카테고리 목록 반환
    설정 파일에서 동적으로 로드 가능
    """
    try:
        # 카테고리 설정 파일 경로 (추후 실제 경로로 수정)
        categories_config_path = "/tmp/categories_config.json"
        
        # 기본 카테고리 (설정 파일이 없을 경우)
        default_categories = {
            "categories": [
                {"id": "tech", "name": "기술", "description": "기술 관련 데이터"},
                {"id": "economy", "name": "경제", "description": "경제 관련 데이터"},
                {"id": "society", "name": "사회", "description": "사회 관련 데이터"},
                {"id": "culture", "name": "문화", "description": "문화 관련 데이터"}
            ],
            "searchable_fields": [
                {"field": "title", "name": "제목", "type": "text"},
                {"field": "content", "name": "내용", "type": "text"},
                {"field": "date", "name": "날짜", "type": "date"},
                {"field": "tags", "name": "태그", "type": "array"}
            ]
        }
        
        if os.path.exists(categories_config_path):
            with open(categories_config_path, 'r', encoding='utf-8') as f:
                categories_config = json.load(f)
        else:
            categories_config = default_categories
        
        return categories_config
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"카테고리 로드 중 오류 발생: {str(e)}")

@app.post("/api/download")
async def request_download(request: DownloadRequest):
    """프론트에서 전달한 데이터를 즉시 Excel로 반환"""
    try:
        if request.filtered_data:
            filtered_data = request.filtered_data
        else:
            filtered_data = await filter_data_by_conditions(request.search_conditions)

        if not filtered_data:
            raise HTTPException(status_code=400, detail="다운로드할 데이터가 없습니다")

        filename = f"datapage_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return _stream_excel_from_records(filtered_data, filename=filename)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"다운로드 요청 처리 실패: {str(e)}")

@app.post("/api/download-search/{category}/{subcategory}")
async def request_search_download(
    category: str,
    subcategory: str,
    request: SearchDownloadRequest,
):
    """검색 조건 기반 Excel 파일을 즉시 스트리밍"""
    try:
        return await _stream_excel_from_query(
            category=category,
            subcategory=subcategory,
            result_type=None,
            keyword=request.keyword,
            search_field=request.search_field,
            filters=request.filters,
            date_from=request.date_from,
            date_to=request.date_to,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"검색 다운로드 처리 실패: {str(e)}")
        raise HTTPException(status_code=500, detail=f"다운로드 요청 처리 실패: {str(e)}")

@app.post("/api/download-search/dataA/{subcategory}")
async def request_search_download_data_a(
    subcategory: str,
    request: SearchDownloadRequest,
):
    return await request_search_download("dataA", subcategory, request)


@app.post("/api/download-search/dataB/{subcategory}")
async def request_search_download_data_b(
    subcategory: str,
    request: SearchDownloadRequest,
):
    return await request_search_download("dataB", subcategory, request)


@app.post("/api/download-search/dataC/{result_type}/{subcategory}")
async def request_search_download_data_c(
    result_type: str,
    subcategory: str,
    request: SearchDownloadRequest,
):
    """DataC 카테고리 Excel 다운로드 (success/failed)"""
    try:
        return await _stream_excel_from_query(
            category="dataC",
            subcategory=subcategory,
            result_type=result_type,
            keyword=request.keyword,
            search_field=request.search_field,
            filters=request.filters,
            date_from=request.date_from,
            date_to=request.date_to,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"DataC 다운로드 처리 실패: {str(e)}")
        raise HTTPException(status_code=500, detail=f"다운로드 요청 처리 실패: {str(e)}")


@app.get("/api/file-info/{category}/{subcategory}")
async def get_file_info(category: str, subcategory: str):
    """
    파일 정보 및 메타데이터 조회
    """
    try:
        data_file_path = get_data_file_path(category, subcategory)
        if not data_file_path:
            raise HTTPException(status_code=404, detail=f"데이터 파일 URL을 찾을 수 없습니다: {category}/{subcategory}")
        
        data_file_str, is_r2_url, is_tabular, file_size_mb = _inspect_data_source(data_file_path)

        total_records_value: Union[int, str] = "unknown"

        if is_tabular:
            from core.duckdb_processor import DuckDBProcessor

            effective_subcategory = normalize_subcategory(subcategory)
            processor = DuckDBProcessor(
                data_file_str,
                category=category,
                subcategory=effective_subcategory
            )

            fields: List[str] = []
            try:
                fields = await asyncio.to_thread(processor.get_available_fields)
                sample_result = await processor.search_streaming(limit=3)
                pagination_info = sample_result.get("pagination", {})

                metadata = {
                    "total_records": pagination_info.get("total_count", 0),
                    "fields": fields,
                    "file_size_mb": round(file_size_mb, 2),
                    "sample_data": sample_result.get("results", [])[:3],
                    "is_large_file": True if is_r2_url or file_size_mb > 50 else False,
                    "data_source": "r2_url" if is_r2_url else "parquet"
                }
            except Exception as e:
                logger.warning(f"DuckDB 메타데이터 조회 실패, 기본값 사용: {e}")
                metadata = {
                    "total_records": 0,
                    "fields": fields if 'fields' in locals() and fields else [],
                    "file_size_mb": round(file_size_mb, 2),
                    "sample_data": [],
                    "is_large_file": True if is_r2_url or file_size_mb > 50 else False,
                    "data_source": "r2_url" if is_r2_url else "parquet",
                    "error": str(e)
                }
            finally:
                processor.close()

        elif file_size_mb > 50:
            from core.large_file_processor import get_large_file_metadata
            metadata = await get_large_file_metadata(data_file_path)
            metadata["is_large_file"] = True
        else:
            with open(data_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            metadata = {
                "file_info": {
                    "size_mb": round(file_size_mb, 2),
                    "modified_time": datetime.now().isoformat()
                },
                "estimated_record_count": len(data.get("data", [])),
                "available_fields": list(data.get("data", [{}])[0].keys()) if data.get("data") else [],
                "is_large_file": False
            }
        
        return metadata
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"파일 정보 조회 실패: {str(e)}")

@app.get("/api/field-samples/{category}/{subcategory}/{field_name}")
async def get_field_samples(category: str, subcategory: str, field_name: str, limit: int = 100):
    """
    특정 필드의 샘플 값들 조회 (필터 옵션 생성용)
    """
    try:
        data_file_path = get_data_file_path(category, subcategory)
        if not data_file_path:
            raise HTTPException(status_code=404, detail=f"데이터 파일 URL을 찾을 수 없습니다: {category}/{subcategory}")
        
        data_file_str, is_r2_url, is_tabular, file_size_mb = _inspect_data_source(data_file_path)

        if is_tabular:
            from core.duckdb_processor import DuckDBProcessor

            effective_subcategory = normalize_subcategory(subcategory)
            processor = DuckDBProcessor(
                data_file_str,
                category=category,
                subcategory=effective_subcategory
            )
            try:
                samples = processor.get_distinct_values(field_name, limit)
            except Exception as e:
                logger.warning(f"DuckDB 필드 샘플 조회 실패: {e}")
                samples = []
            finally:
                processor.close()
        elif file_size_mb > 50:
            processor = get_processor(data_file_path)
            samples = await processor.get_field_samples(field_name, limit)
        else:
            with open(data_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            samples = set()
            for item in data.get("data", []):
                if field_name in item and item[field_name] is not None:
                    samples.add(item[field_name])
                    if len(samples) >= limit:
                        break
            samples = sorted(list(samples))
        
        return {
            "field_name": field_name,
            "sample_count": len(samples),
            "samples": samples[:limit]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"필드 샘플 조회 실패: {str(e)}")

@app.post("/api/clear-cache")
async def clear_processor_cache():
    """
    대용량 파일 프로세서 캐시 클리어
    """
    try:
        from core.large_file_processor import clear_all_processors
        clear_all_processors()
        return {"message": "캐시가 성공적으로 클리어되었습니다"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"캐시 클리어 실패: {str(e)}")

# ====================================
# 표시 설정 관리 API 엔드포인트들
# ====================================

@app.get("/api/config/{category}/{subcategory}")
async def get_display_config(category: str, subcategory: str):
    """
    카테고리별 표시 설정 조회
    """
    try:
        # get_config는 이제 항상 설정을 반환함 (None 반환 없음)
        config = display_config_manager.get_config(category, subcategory)
        return display_config_manager.export_client_config(category, subcategory)
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"설정 조회 실패: {str(e)}")

@app.put("/api/config/{category}/{subcategory}")
async def update_display_config(category: str, subcategory: str, config: CategoryDisplayConfig):
    """
    카테고리별 표시 설정 업데이트
    """
    try:
        display_config_manager.save_config(category, subcategory, config)
        
        return {
            "message": "설정이 성공적으로 업데이트되었습니다",
            "config": display_config_manager.export_client_config(category, subcategory)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"설정 업데이트 실패: {str(e)}")

@app.patch("/api/config/{category}/{subcategory}")
async def patch_display_config(category: str, subcategory: str, updates: Dict[str, Any]):
    """
    카테고리별 표시 설정 부분 업데이트
    """
    try:
        updated_config = display_config_manager.update_config(category, subcategory, updates)
        
        return {
            "message": "설정이 성공적으로 업데이트되었습니다",
            "config": display_config_manager.export_client_config(category, subcategory)
        }
        
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"설정 업데이트 실패: {str(e)}")

@app.delete("/api/config/{category}/{subcategory}")
async def delete_display_config(category: str, subcategory: str):
    """
    카테고리별 표시 설정 삭제 (기본 설정으로 재생성됨)
    """
    try:
        display_config_manager.delete_config(category, subcategory)
        
        return {"message": "설정이 성공적으로 삭제되었습니다"}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"설정 삭제 실패: {str(e)}")

@app.get("/api/configs")
async def list_all_configs():
    """
    모든 카테고리의 표시 설정 목록 조회
    """
    try:
        configs = display_config_manager.list_configs()
        
        result = {}
        for key, config in configs.items():
            category, subcategory = key.split('/', 1)
            result[key] = {
                "category": category,
                "subcategory": subcategory,
                "display_name": config.display_name,
                "description": config.description,
                "display_fields_count": len(config.display_fields),
                "download_fields_count": len(config.download_fields),
                "search_fields_count": len(config.search_fields)
            }
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"설정 목록 조회 실패: {str(e)}")

@app.get("/api/config/{category}/{subcategory}/validate")
async def validate_config_against_data(category: str, subcategory: str):
    """
    설정이 실제 데이터와 일치하는지 검증
    """
    try:
        validation_result = display_config_manager.validate_fields_against_data(category, subcategory)
        return validation_result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"설정 검증 실패: {str(e)}")

@app.get("/api/config/{category}/{subcategory}/preview")
async def preview_config(category: str, subcategory: str, limit: int = 5):
    """
    설정 기반 데이터 미리보기
    """
    try:
        # 설정 조회 (get_config는 항상 설정을 반환함)
        config = display_config_manager.get_config(category, subcategory)
        
        # 실제 데이터 URL 가져오기
        data_file_path = get_data_file_path(category, subcategory)
        if not data_file_path:
            raise HTTPException(status_code=404, detail=f"데이터 파일 URL을 찾을 수 없습니다: {category}/{subcategory}")
        
        data_file_str, is_r2_url, is_tabular, file_size_mb = _inspect_data_source(data_file_path)

        if is_tabular:
            search_result = await duckdb_search_large_file(
                file_path=data_file_str,
                keyword=None,
                search_field="all",
                limit=limit,
                page=1,
                filters=None,
                category=category,
                subcategory=normalize_subcategory(subcategory)
            )
            preview_data = search_result.get("results", [])
        elif file_size_mb > 50:
            search_result = await stream_search_large_file(
                file_path=data_file_path,
                keyword=None,
                search_field="all",
                filters=None,
                limit=limit,
                offset=0
            )
            preview_data = search_result.get("results", [])
        else:
            # 일반 파일 처리
            with open(data_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            preview_data = data.get("data", [])[:limit]
        
        # 설정 기반으로 필터링된 데이터 반환
        filtered_preview = []
        display_field_names = [field.field for field in config.display_fields]
        
        for item in preview_data:
            # resultData 구조 처리
            if "resultData" in item:
                item_data = item["resultData"]
            else:
                item_data = item
            
            # 표시 필드만 추출
            filtered_item = {}
            for field in config.display_fields:
                filtered_item[field.field] = item_data.get(field.field, "")
            
            filtered_preview.append(filtered_item)
        
        return {
            "config": display_config_manager.export_client_config(category, subcategory),
            "preview_data": filtered_preview,
            "total_preview_count": len(filtered_preview)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"미리보기 생성 실패: {str(e)}")

@app.post("/api/config/{category}/{subcategory}/generate")
async def generate_config_from_data(category: str, subcategory: str):
    """
    실제 데이터를 분석하여 자동으로 설정 생성
    """
    try:
        # 기존 설정 삭제 (있다면)
        try:
            display_config_manager.delete_config(category, subcategory)
        except:
            pass
        
        # 새 설정 생성 (자동 분석)
        new_config = display_config_manager.get_config(category, subcategory)
        
        return {
            "message": "설정이 자동으로 생성되었습니다",
            "config": display_config_manager.export_client_config(category, subcategory)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"자동 설정 생성 실패: {str(e)}")

@app.get("/api/settings/{category}/{subcategory}")
async def get_field_settings(category: str, subcategory: str):
    """
    필드 설정 조회 (클라이언트용) - 2-parameter
    """
    try:
        # get_config는 이제 항상 설정을 반환함 (None 반환 없음)
        config = display_config_manager.get_config(category, subcategory)
        return display_config_manager.export_client_config(category, subcategory)
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"설정 조회 실패: {str(e)}")

@app.get("/api/settings/{category}/{result_type}/{subcategory}")
async def get_field_settings_3param(category: str, result_type: str, subcategory: str):
    """
    필드 설정 조회 (클라이언트용) - 3-parameter for dataC
    """
    try:
        # dataC의 경우: category=dataC, result_type=success, subcategory=safetykorea
        # DisplayConfigManager가 이제 dataC/success/safetykorea 3-level key를 지원함
        
        if category == 'dataC':
            # dataC의 경우: dataC/success/safetykorea key 사용
            config = display_config_manager.get_config(f"{category}/{result_type}", subcategory)
            return display_config_manager.export_client_config(f"{category}/{result_type}", subcategory)
        else:
            # 다른 카테고리는 기존 방식 사용
            config = display_config_manager.get_config(category, f"{result_type}/{subcategory}")
            return display_config_manager.export_client_config(category, f"{result_type}/{subcategory}")
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"설정 조회 실패 (3-param): {str(e)}")

@app.put("/api/settings/{category}/{subcategory}")
async def update_field_settings(category: str, subcategory: str, settings: Dict[str, Any]):
    """
    필드 설정 업데이트
    """
    try:
        # 설정 변환 및 업데이트
        config_data = {
            "display_name": settings.get("displayName", ""),
            "description": settings.get("description", ""),
            "display_fields": [],
            "download_fields": settings.get("downloadFields", []),
            "search_fields": [],
            "default_sort_field": settings.get("defaultSort", {}).get("field"),
            "default_sort_order": settings.get("defaultSort", {}).get("order", "asc"),
            "items_per_page": settings.get("pagination", {}).get("itemsPerPage", 20),
            "enable_export": settings.get("features", {}).get("enableExport", True),
            "show_summary": settings.get("features", {}).get("showSummary", True),
            "show_pagination": settings.get("features", {}).get("showPagination", True)
        }
        
        # 표시 필드 변환
        for field_data in settings.get("displayFields", []):
            display_field = DisplayField(**field_data)
            config_data["display_fields"].append(display_field)
        
        # 검색 필드 변환
        for field_data in settings.get("searchFields", []):
            search_field = SearchField(**field_data)
            config_data["search_fields"].append(search_field)
        
        # 설정 저장
        config = CategoryDisplayConfig(**config_data)
        display_config_manager.save_config(category, subcategory, config)
        
        return {
            "message": "설정이 성공적으로 업데이트되었습니다",
            "config": display_config_manager.export_client_config(category, subcategory)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"설정 업데이트 실패: {str(e)}")

@app.get("/api/settings/preview/{category}/{subcategory}")
async def get_settings_preview(category: str, subcategory: str, limit: int = 5):
    """
    설정 미리보기 (실제 데이터와 함께)
    """
    try:
        # 기존 미리보기 API 활용
        return await preview_config(category, subcategory, limit)
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"미리보기 생성 실패: {str(e)}")

@app.get("/api/field-info/{category}/{subcategory}")
async def get_field_information(category: str, subcategory: str):
    """
    데이터 파일에서 필드 정보 추출 (설정 UI용)
    """
    try:
        data_file_path = get_data_file_path(category, subcategory)
        if not data_file_path:
            raise HTTPException(status_code=404, detail=f"데이터 파일 URL을 찾을 수 없습니다: {category}/{subcategory}")
        
        data_file_str, is_r2_url, is_tabular, file_size_mb = _inspect_data_source(data_file_path)

        if is_tabular:
            from core.duckdb_processor import DuckDBProcessor

            effective_subcategory = normalize_subcategory(subcategory)
            processor = DuckDBProcessor(
                data_file_str,
                category=category,
                subcategory=effective_subcategory
            )

            try:
                available_fields = processor.get_available_fields()
                sample_result = await processor.search_streaming(limit=3)
                sample_data = sample_result.get("results", [])
                total_records_value = sample_result.get("pagination", {}).get("total_count", "unknown")
            finally:
                processor.close()
        elif file_size_mb > 50:
            from core.large_file_processor import get_large_file_metadata
            metadata = await get_large_file_metadata(data_file_path)
            available_fields = metadata.get("available_fields", [])
            sample_data = []
            total_records_value = metadata.get("estimated_record_count", "unknown")
        else:
            # JSON 구조 처리
            with open(data_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if data.get("data") and len(data["data"]) > 0:
                first_record = data["data"][0]
                # resultData 구조 처리
                if "resultData" in first_record:
                    first_record = first_record["resultData"]
                available_fields = list(first_record.keys())
                sample_data = data["data"][:3]  # 샘플 3개
            else:
                available_fields = []
                sample_data = []
            total_records_value = len(data.get("data", [])) if data.get("data") else 0
        
        # 필드 타입 추론
        field_types = {}
        for field in available_fields:
            field_types[field] = infer_field_type(sample_data, field)
        
        return {
            "available_fields": available_fields,
            "field_types": field_types,
            "sample_data": sample_data[:3],  # 샘플 3개만
            "total_records": total_records_value,
            "is_large_file": file_size_mb > 50 or is_r2_url,
            "file_size_mb": round(file_size_mb, 2)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"필드 정보 조회 실패: {str(e)}")

# 유틸리티 함수들
def get_category_name(category_id: str, categories: List[Dict]) -> str:
    """카테고리 ID로 이름 찾기"""
    for cat in categories:
        if cat.get("id") == category_id:
            return cat.get("name", category_id)
    return category_id

def get_korean_field_mapping(category: str, subcategory: str) -> Dict[str, str]:
    """카테고리/서브카테고리에 따른 영어->한글 필드명 매핑 반환"""
    try:
        normalized_subcategory = normalize_subcategory(subcategory)
        config_path = Path(__file__).parent.parent / "config" / "field_settings.json"
        with open(config_path, 'r', encoding='utf-8') as f:
            field_settings = json.load(f)

        # 카테고리별 설정 확인
        if category in field_settings and normalized_subcategory in field_settings[category]:
            display_fields = field_settings[category][normalized_subcategory].get("display_fields", [])
            # field → name 매핑 생성
            field_mapping = {}
            for field_config in display_fields:
                english_field = field_config.get("field")
                korean_name = field_config.get("name")
                if english_field and korean_name:
                    field_mapping[english_field] = korean_name
            
            return field_mapping
    except Exception as e:
        logger.error(f"필드명 매핑 로드 실패: {str(e)}")
    
    return {}  # 실패 시 빈 딕셔너리 반환

def get_download_fields(category: str, subcategory: str, result_type: Optional[str] = None) -> List[str]:
    """카테고리/서브카테고리에 따른 download_fields 목록 반환"""
    try:
        normalized_subcategory = normalize_subcategory(subcategory)
        config_path = Path(__file__).parent.parent / "config" / "field_settings.json"
        with open(config_path, 'r', encoding='utf-8') as f:
            field_settings = json.load(f)

        category_config = field_settings.get(category, {})

        if category == "dataC":
            bucket_key = result_type or "success"
            category_config = category_config.get(bucket_key, {})

        if normalized_subcategory in category_config:
            download_fields = category_config[normalized_subcategory].get("download_fields", [])
            return download_fields
    except Exception as e:
        logger.error(f"download_fields 로드 실패: {str(e)}")

    return []  # 실패 시 빈 리스트 반환

def filter_data_by_download_fields(data: List[Dict], download_fields: List[str]) -> List[Dict]:
    """데이터를 download_fields에 지정된 필드만 포함하도록 필터링"""
    if not download_fields or not data:
        return data

    filtered_data = []
    for item in data:
        filtered_item = {}
        for field in download_fields:
            if field in item:
                filtered_item[field] = item[field]
        filtered_data.append(filtered_item)

    return filtered_data


def _normalize_excel_value(value: Any) -> Any:
    """엑셀 출력용 값 전처리"""
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        try:
            return json.dumps(value, ensure_ascii=False)
        except Exception:
            return str(value)
    return value


async def _stream_excel_from_query(
    *,
    category: str,
    subcategory: str,
    result_type: Optional[str],
    keyword: Optional[str],
    search_field: Optional[str],
    filters: Optional[Dict[str, Any]],
    date_from: Optional[str],
    date_to: Optional[str],
) -> StreamingResponse:
    """DuckDB 조회 결과를 즉시 Excel로 스트리밍"""

    # 데이터 파일 경로 확인
    if category == "dataC" and result_type:
        data_file_path = get_data_file_path_c(category, result_type, subcategory)
    else:
        data_file_path = get_data_file_path(category, subcategory)

    effective_subcategory = normalize_subcategory(subcategory)

    korean_field_mapping = get_korean_field_mapping(category, subcategory)
    download_fields = get_download_fields(category, subcategory, result_type)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        temp_path = Path(tmp.name)

    total_records = 0
    header_written = False
    selected_columns: List[str] = []

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title="검색 결과")

    def handle_chunk(chunk: List[Dict[str, Any]], total_processed: int):
        nonlocal header_written, selected_columns, total_records
        if not chunk:
            return

        if not header_written:
            if download_fields:
                selected_columns = download_fields
            else:
                selected_columns = list(chunk[0].keys())

            header_labels = [korean_field_mapping.get(col, col) for col in selected_columns]
            worksheet.append(header_labels)
            header_written = True

        for record in chunk:
            row = [_normalize_excel_value(record.get(col)) for col in selected_columns]
            worksheet.append(row)

        total_records = total_processed

    query_result = await duckdb_search_large_file(
        file_path=str(data_file_path),
        keyword=keyword,
        search_field=search_field,
        limit=None,
        page=1,
        filters=filters,
        category=category,
        subcategory=effective_subcategory,
        result_type=result_type,
        collect_results=False,
        chunk_callback=handle_chunk,
        chunk_size=1000,
        required_fields=download_fields
    )

    if query_result.get("error"):
        raise HTTPException(status_code=500, detail=f"검색 실패: {query_result.get('message')}")

    if not header_written:
        try:
            temp_path.unlink(missing_ok=True)
        except Exception:
            pass
        raise HTTPException(status_code=404, detail="검색 결과가 없습니다.")

    workbook.save(temp_path)
    workbook.close()

    filename = f"datapage_{effective_subcategory}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    def file_iterator():
        try:
            with temp_path.open('rb') as f:
                for chunk in iter(lambda: f.read(64 * 1024), b''):
                    yield chunk
        finally:
            try:
                temp_path.unlink(missing_ok=True)
            except Exception:
                pass

    response = StreamingResponse(
        file_iterator(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
    response.headers["X-Total-Count"] = str(query_result.get("pagination", {}).get("total_count", total_records))
    return response


def _stream_excel_from_records(records: List[Dict[str, Any]], *, filename: Optional[str] = None) -> StreamingResponse:
    """프론트에서 전달된 소량 데이터를 즉시 Excel로 반환"""
    if not records:
        raise HTTPException(status_code=400, detail="다운로드할 데이터가 없습니다")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        temp_path = Path(tmp.name)

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title="검색 결과")

    headers = list(records[0].keys())
    worksheet.append(headers)

    for record in records:
        row = [_normalize_excel_value(record.get(col)) for col in headers]
        worksheet.append(row)

    workbook.save(temp_path)
    workbook.close()

    if not filename:
        filename = f"datapage_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    def file_iterator():
        try:
            with temp_path.open('rb') as f:
                for chunk in iter(lambda: f.read(64 * 1024), b''):
                    yield chunk
        finally:
            try:
                temp_path.unlink(missing_ok=True)
            except Exception:
                pass

    response = StreamingResponse(
        file_iterator(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
    response.headers["X-Total-Count"] = str(len(records))
    return response

def get_data_file_path_c(category: str, result_type: str, subcategory: str, prefer_r2: bool = False) -> str:
    """DataC 카테고리용 3-parameter 데이터 파일 경로 생성 (DATA_MODE 환경변수 기반 듀얼 모드)"""

    normalized_subcategory = normalize_subcategory(subcategory)
    if normalized_subcategory != subcategory:
        logger.info(
            f"서브카테고리 정규화: {category}/{result_type}/{subcategory} → {normalized_subcategory}"
        )
        subcategory = normalized_subcategory

    # 🎯 DATA_MODE 환경변수로 모드 결정
    data_mode = os.getenv("DATA_MODE", "full").lower()

    # Vercel Blob URL 매핑 (DataC 2025년 필터링된 데이터) - 환경변수 사용
    blob_env_mapping_c = {
        # DataC Success 매핑
        ("success", "safetykorea"): "BLOB_URL_DATAC_SUCCESS_1_SAFETYKOREA",
        ("success", "wadiz-makers"): "BLOB_URL_DATAC_SUCCESS_2_WADIZ",
        ("success", "efficiency-rating"): "BLOB_URL_DATAC_SUCCESS_3_EFFICIENCY",
        ("success", "high-efficiency"): "BLOB_URL_DATAC_SUCCESS_4_HIGH_EFFICIENCY",
        ("success", "standby-power"): "BLOB_URL_DATAC_SUCCESS_5_STANDBY_POWER",
        ("success", "approval"): "BLOB_URL_DATAC_SUCCESS_6_APPROVAL",
        ("success", "declaration-details"): "BLOB_URL_DATAC_SUCCESS_7_DECLARE",
        ("success", "kwtc"): "BLOB_URL_DATAC_SUCCESS_8_KWTC",
        ("success", "recall"): "BLOB_URL_DATAC_SUCCESS_9_RECALL",
        ("success", "safetykoreachild"): "BLOB_URL_DATAC_SUCCESS_10_SAFETYKOREACHILD",
        ("success", "rra-cert"): "BLOB_URL_DATAC_SUCCESS_11_RRA_CERT",
        ("success", "rra-self-cert"): "BLOB_URL_DATAC_SUCCESS_12_RRA_SELF_CERT",
        ("success", "safetykoreahome"): "BLOB_URL_DATAC_SUCCESS_13_SAFETYKOREAHOME",

        # DataC Failed 매핑
        ("failed", "safetykorea"): "BLOB_URL_DATAC_FAILED_1_SAFETYKOREA",
        ("failed", "wadiz-makers"): "BLOB_URL_DATAC_FAILED_2_WADIZ",
        ("failed", "efficiency-rating"): "BLOB_URL_DATAC_FAILED_3_EFFICIENCY",
        ("failed", "high-efficiency"): "BLOB_URL_DATAC_FAILED_4_HIGH_EFFICIENCY",
        ("failed", "standby-power"): "BLOB_URL_DATAC_FAILED_5_STANDBY_POWER",
        ("failed", "approval"): "BLOB_URL_DATAC_FAILED_6_APPROVAL",
        ("failed", "declaration-details"): "BLOB_URL_DATAC_FAILED_7_DECLARE",
        ("failed", "kwtc"): "BLOB_URL_DATAC_FAILED_8_KWTC",
        ("failed", "recall"): "BLOB_URL_DATAC_FAILED_9_RECALL",
        ("failed", "safetykoreachild"): "BLOB_URL_DATAC_FAILED_10_SAFETYKOREACHILD",
        ("failed", "rra-cert"): "BLOB_URL_DATAC_FAILED_11_RRA_CERT",
        ("failed", "rra-self-cert"): "BLOB_URL_DATAC_FAILED_12_RRA_SELF_CERT",
        ("failed", "safetykoreahome"): "BLOB_URL_DATAC_FAILED_13_SAFETYKOREAHOME"
    }

    # 로컬 parquet 파일 경로 매핑 (2025년 필터링된 데이터)
    local_file_mapping = {
        "safetykorea": "./parquet/1_safetykorea_flattened.parquet",
        "kwtc": "./parquet/8_kwtc_flattened.parquet",
        "rra-cert": "./parquet/11_rra_cert_flattened.parquet",
        "rra-self-cert": "./parquet/12_rra_self_cert_flattened.parquet",
        "efficiency-rating": "./parquet/3_efficiency_flattened.parquet",
        "high-efficiency": "./parquet/4_high_efficiency_flattened.parquet",
        "standby-power": "./parquet/5_standby_power_flattened.parquet",
        "approval": "./parquet/6_approval_flattened.parquet",
        "declaration-details": "./parquet/7_declare_flattened.parquet",
        "recall": "./parquet/9_recall_flattened.parquet",
        "safetykoreachild": "./parquet/10_safetykoreachild_flattened.parquet",
        "safetykoreahome": "./parquet/13_safetykoreahome_flattened.parquet",
        "wadiz-makers": "./parquet/2_wadiz_flattened.parquet",
    }

    # 🟢 2025년 모드: Vercel Blob URL 우선 사용 (성능 최적화)
    if not prefer_r2 and data_mode == "2025":
        # 1. Vercel Blob URL 사용 (환경변수에서)
        blob_env_var = blob_env_mapping_c.get((result_type, subcategory))
        if blob_env_var:
            blob_url = os.getenv(blob_env_var)
            if blob_url:
                logger.info(f"2025년 모드 (Blob): {category}/{result_type}/{subcategory} → {blob_url}")
                return blob_url
            else:
                logger.warning(f"Blob 환경변수 없음: {blob_env_var}, prefetch로 fallback")

        # 2. Prefetch 시스템 fallback
        prefetched_path = get_prefetched_blob_path(category, subcategory, result_type)
        if prefetched_path:
            logger.info(f"2025년 모드 (Blob-prefetch): {category}/{result_type}/{subcategory} → {prefetched_path}")
            return prefetched_path

        local_parquet_path = local_file_mapping.get(subcategory, "./parquet/1_safetykorea_flattened.parquet")

        # Vercel 환경에서 절대경로도 시도
        if not os.path.exists(local_parquet_path):
            # 작업 디렉토리 기준 절대경로 시도
            abs_parquet_path = os.path.abspath(local_parquet_path)
            if os.path.exists(abs_parquet_path):
                logger.info(f"2025년 모드 (DataC-절대경로): {category}/{result_type}/{subcategory} → {abs_parquet_path}")
                return abs_parquet_path

            # Project/ 하위 경로 시도
            project_parquet_path = f"Project/{local_parquet_path}"
            if os.path.exists(project_parquet_path):
                logger.info(f"2025년 모드 (DataC-Project/): {category}/{result_type}/{subcategory} → {project_parquet_path}")
                return project_parquet_path

        if os.path.exists(local_parquet_path):
            logger.info(f"2025년 모드 (DataC): {category}/{result_type}/{subcategory} → {local_parquet_path}")
            return local_parquet_path
        else:
            # 2025년 모드에서 파일을 찾을 수 없으면 R2로 fallback
            logger.warning(f"2025년 모드 (DataC): 파일 없음 {local_parquet_path}, R2 모드로 fallback")
            # R2 모드로 처리하도록 아래 R2 로직으로 진행

    # 🔵 전체 데이터 모드: R2 URL 사용 (기본값, 프로덕션)
    r2_url_mapping = {
        # DataC Success 구조 매핑
        ("dataC", "success", "safetykorea"): os.getenv("R2_URL_DATAC_SUCCESS_1_SAFETYKOREA"),
        ("dataC", "success", "wadiz-makers"): os.getenv("R2_URL_DATAC_SUCCESS_2_WADIZ"),
        ("dataC", "success", "efficiency-rating"): os.getenv("R2_URL_DATAC_SUCCESS_3_EFFICIENCY"),
        ("dataC", "success", "high-efficiency"): os.getenv("R2_URL_DATAC_SUCCESS_4_HIGH_EFFICIENCY"),
        ("dataC", "success", "standby-power"): os.getenv("R2_URL_DATAC_SUCCESS_5_STANDBY_POWER"),
        ("dataC", "success", "approval"): os.getenv("R2_URL_DATAC_SUCCESS_6_APPROVAL"),
        ("dataC", "success", "declaration-details"): os.getenv("R2_URL_DATAC_SUCCESS_7_DECLARE"),
        ("dataC", "success", "kwtc"): os.getenv("R2_URL_DATAC_SUCCESS_8_KWTC"),
        ("dataC", "success", "recall"): os.getenv("R2_URL_DATAC_SUCCESS_9_RECALL"),
        ("dataC", "success", "safetykoreachild"): os.getenv("R2_URL_DATAC_SUCCESS_10_SAFETYKOREACHILD"),
        ("dataC", "success", "safetykoreahome"): os.getenv("R2_URL_DATAC_SUCCESS_13_SAFETYKOREAHOME"),
        ("dataC", "success", "rra-cert"): os.getenv("R2_URL_DATAC_SUCCESS_11_RRA_CERT"),
        ("dataC", "success", "rra-self-cert"): os.getenv("R2_URL_DATAC_SUCCESS_12_RRA_SELF_CERT"),

        # DataC Failed 구조 매핑
        ("dataC", "failed", "safetykorea"): os.getenv("R2_URL_DATAC_FAILED_1_SAFETYKOREA"),
        ("dataC", "failed", "wadiz-makers"): os.getenv("R2_URL_DATAC_FAILED_2_WADIZ"),
        ("dataC", "failed", "efficiency-rating"): os.getenv("R2_URL_DATAC_FAILED_3_EFFICIENCY"),
        ("dataC", "failed", "high-efficiency"): os.getenv("R2_URL_DATAC_FAILED_4_HIGH_EFFICIENCY"),
        ("dataC", "failed", "standby-power"): os.getenv("R2_URL_DATAC_FAILED_5_STANDBY_POWER"),
        ("dataC", "failed", "approval"): os.getenv("R2_URL_DATAC_FAILED_6_APPROVAL"),
        ("dataC", "failed", "declaration-details"): os.getenv("R2_URL_DATAC_FAILED_7_DECLARE"),
        ("dataC", "failed", "kwtc"): os.getenv("R2_URL_DATAC_FAILED_8_KWTC"),
        ("dataC", "failed", "recall"): os.getenv("R2_URL_DATAC_FAILED_9_RECALL"),
        ("dataC", "failed", "safetykoreachild"): os.getenv("R2_URL_DATAC_FAILED_10_SAFETYKOREACHILD"),
        ("dataC", "failed", "safetykoreahome"): os.getenv("R2_URL_DATAC_FAILED_13_SAFETYKOREAHOME"),
        ("dataC", "failed", "rra-cert"): os.getenv("R2_URL_DATAC_FAILED_11_RRA_CERT"),
        ("dataC", "failed", "rra-self-cert"): os.getenv("R2_URL_DATAC_FAILED_12_RRA_SELF_CERT"),
    }

    r2_url = r2_url_mapping.get((category, result_type, subcategory))

    # 로컬 개발 환경 fallback (VERCEL 환경변수 없을 때)
    if not r2_url and os.getenv("VERCEL") is None:
        local_parquet_path = local_file_mapping.get(subcategory, "./parquet/1_safetykorea_flattened.parquet")
        if os.path.exists(local_parquet_path):
            logger.info(f"로컬 개발 환경 (DataC): {category}/{result_type}/{subcategory} → {local_parquet_path}")
            return local_parquet_path
        else:
            logger.info(f"로컬 개발 환경 (DataC): 기본 파일 사용")
            return "./parquet/1_safetykorea_flattened.parquet"

    if not r2_url:
        # 환경변수가 없으면 로컬 파일 후보로 마지막 시도
        fallback_candidates = []
        local_candidate = local_file_mapping.get(subcategory)
        if local_candidate:
            fallback_candidates.extend([
                local_candidate,
                os.path.abspath(local_candidate),
                f"Project/{local_candidate}"
            ])

        for candidate in fallback_candidates:
            if candidate and os.path.exists(candidate):
                logger.warning(
                    f"R2 URL 누락: {category}/{result_type}/{subcategory} - 로컬 파일로 대체 ({candidate})"
                )
                return candidate

        raise ValueError(f"R2 URL not found for {category}/{result_type}/{subcategory}. Check environment variables.")

    logger.info(f"전체 데이터 모드 (DataC): {category}/{result_type}/{subcategory} → R2")
    return r2_url

def get_data_file_path(category: str, subcategory: str, prefer_r2: bool = False) -> str:
    """카테고리와 서브카테고리로 데이터 파일 경로 생성 (DATA_MODE 환경변수 기반 듀얼 모드)"""

    normalized_subcategory = normalize_subcategory(subcategory)
    if normalized_subcategory != subcategory:
        logger.info(
            f"서브카테고리 정규화: {category}/{subcategory} → {normalized_subcategory}"
        )
        subcategory = normalized_subcategory

    # 🎯 DATA_MODE 환경변수로 모드 결정
    data_mode = os.getenv("DATA_MODE", "full").lower()

    # Vercel Blob URL 매핑 (2025년 필터링된 데이터) - 환경변수 사용
    blob_env_mapping = {
        # DataA 매핑 (12개)
        "safetykorea": "BLOB_URL_DATAA_1_SAFETYKOREA",
        "efficiency-rating": "BLOB_URL_DATAA_3_EFFICIENCY",
        "high-efficiency": "BLOB_URL_DATAA_4_HIGH_EFFICIENCY",
        "standby-power": "BLOB_URL_DATAA_5_STANDBY_POWER",
        "approval": "BLOB_URL_DATAA_6_APPROVAL",
        "declaration-details": "BLOB_URL_DATAA_7_DECLARE",
        "kwtc": "BLOB_URL_DATAA_8_KWTC",
        "recall": "BLOB_URL_DATAA_9_RECALL",
        "safetykoreachild": "BLOB_URL_DATAA_10_SAFETYKOREACHILD",
        "rra-cert": "BLOB_URL_DATAA_11_RRA_CERT",
        "rra-self-cert": "BLOB_URL_DATAA_12_RRA_SELF_CERT",
        "safetykoreahome": "BLOB_URL_DATAA_13_SAFETYKOREAHOME",

        # DataB 매핑 (1개)
        "wadiz-makers": "BLOB_URL_DATAB_2_WADIZ",
    }

    # 로컬 parquet 파일 경로 매핑 (2025년 필터링된 데이터)
    local_file_mapping = {
        # DataA 매핑
        "safetykorea": "./parquet/1_safetykorea_flattened.parquet",
        "kwtc": "./parquet/8_kwtc_flattened.parquet",
        "rra-cert": "./parquet/11_rra_cert_flattened.parquet",     # RRA 인증
        "rra-self-cert": "./parquet/12_rra_self_cert_flattened.parquet",  # RRA 자기적합성
        "efficiency-rating": "./parquet/3_efficiency_flattened.parquet",
        "high-efficiency": "./parquet/4_high_efficiency_flattened.parquet",
        "standby-power": "./parquet/5_standby_power_flattened.parquet",
        "approval": "./parquet/6_approval_flattened.parquet",           # 승인정보
        "declaration-details": "./parquet/7_declare_flattened.parquet",
        "recall": "./parquet/9_recall_flattened.parquet",
        "safetykoreachild": "./parquet/10_safetykoreachild_flattened.parquet",
        "safetykoreahome": "./parquet/13_safetykoreahome_flattened.parquet",
        # DataB 매핑
        "wadiz-makers": "./parquet/2_wadiz_flattened.parquet",  # 와디즈 메이커
    }

    # 🟢 2025년 모드: Vercel Blob URL 우선 사용 (성능 최적화)
    if not prefer_r2 and data_mode == "2025":
        prefetched_path = get_prefetched_blob_path(category, subcategory)
        if prefetched_path:
            logger.info(f"2025년 모드 (Blob-prefetch): {category}/{subcategory} → {prefetched_path}")
            return prefetched_path

        # 1. Vercel Blob URL 사용 (환경변수에서)
        blob_env_var = blob_env_mapping.get(subcategory)
        if blob_env_var:
            blob_url = os.getenv(blob_env_var)
            if blob_url:
                logger.info(f"2025년 모드 (Blob): {category}/{subcategory} → {blob_url}")
                return blob_url
            else:
                logger.warning(f"Blob 환경변수 없음: {blob_env_var}, 로컬 파일로 fallback")

        # 2. 로컬 파일 fallback
        local_parquet_path = local_file_mapping.get(subcategory, "./parquet/1_safetykorea_flattened.parquet")

        # Vercel 환경에서 절대경로도 시도
        if not os.path.exists(local_parquet_path):
            # 작업 디렉토리 기준 절대경로 시도
            abs_parquet_path = os.path.abspath(local_parquet_path)
            if os.path.exists(abs_parquet_path):
                logger.info(f"2025년 모드 (절대경로): {category}/{subcategory} → {abs_parquet_path}")
                return abs_parquet_path

            # Project/ 하위 경로 시도
            project_parquet_path = f"Project/{local_parquet_path}"
            if os.path.exists(project_parquet_path):
                logger.info(f"2025년 모드 (Project/): {category}/{subcategory} → {project_parquet_path}")
                return project_parquet_path

        if os.path.exists(local_parquet_path):
            logger.info(f"2025년 모드: {category}/{subcategory} → {local_parquet_path}")
            return local_parquet_path
        else:
            # 2025년 모드에서 파일을 찾을 수 없으면 R2로 fallback
            logger.warning(f"2025년 모드: 파일 없음 {local_parquet_path}, R2 모드로 fallback")
            # R2 모드로 처리하도록 data_mode 변경하지 않고 아래 R2 로직으로 진행

    # 🔵 전체 데이터 모드: R2 URL 사용 (기본값, 프로덕션)
    r2_url_mapping = {
        # DataA 구조 매핑 - 새로운 환경변수 이름 사용
        ("dataA", "safetykorea"): os.getenv("R2_URL_DATAA_1_SAFETYKOREA"),
        ("dataA", "kwtc"): os.getenv("R2_URL_DATAA_8_KWTC"),
        ("dataA", "rra-cert"): os.getenv("R2_URL_DATAA_11_RRA_CERT"),     # RRA 인증
        ("dataA", "rra-self-cert"): os.getenv("R2_URL_DATAA_12_RRA_SELF_CERT"),  # RRA 자기적합성
        ("dataA", "efficiency-rating"): os.getenv("R2_URL_DATAA_3_EFFICIENCY"), # 효율등급
        ("dataA", "high-efficiency"): os.getenv("R2_URL_DATAA_4_HIGH_EFFICIENCY"),        # 고효율기기
        ("dataA", "standby-power"): os.getenv("R2_URL_DATAA_5_STANDBY_POWER"),      # 대기전력
        ("dataA", "approval"): os.getenv("R2_URL_DATAA_6_APPROVAL"),           # 승인정보
        ("dataA", "declaration-details"): os.getenv("R2_URL_DATAA_7_DECLARE"),     # 신고정보
        ("dataA", "recall"): os.getenv("R2_URL_DATAA_9_RECALL"),               # 리콜정보(국내)
        ("dataA", "safetykoreachild"): os.getenv("R2_URL_DATAA_10_SAFETYKOREACHILD"),  # 어린이용품 인증정보
        ("dataA", "safetykoreahome"): os.getenv("R2_URL_DATAA_13_SAFETYKOREAHOME"),  # 생활용품

        # DataB 구조 매핑 - DataA와 동일한 파일 사용
        ("dataB", "wadiz-makers"): os.getenv("R2_URL_DATAB_2_WADIZ"),              # 와디즈 메이커

        # DataC Success 구조 매핑
        ("dataC", "success", "safetykorea"): os.getenv("R2_URL_DATAC_SUCCESS_1_SAFETYKOREA"),
        ("dataC", "success", "wadiz-makers"): os.getenv("R2_URL_DATAC_SUCCESS_2_WADIZ"),
        ("dataC", "success", "efficiency-rating"): os.getenv("R2_URL_DATAC_SUCCESS_3_EFFICIENCY"),
        ("dataC", "success", "high-efficiency"): os.getenv("R2_URL_DATAC_SUCCESS_4_HIGH_EFFICIENCY"),
        ("dataC", "success", "standby-power"): os.getenv("R2_URL_DATAC_SUCCESS_5_STANDBY_POWER"),
        ("dataC", "success", "approval"): os.getenv("R2_URL_DATAC_SUCCESS_6_APPROVAL"),
        ("dataC", "success", "declaration-details"): os.getenv("R2_URL_DATAC_SUCCESS_7_DECLARE"),
        ("dataC", "success", "kwtc"): os.getenv("R2_URL_DATAC_SUCCESS_8_KWTC"),
        ("dataC", "success", "recall"): os.getenv("R2_URL_DATAC_SUCCESS_9_RECALL"),
        ("dataC", "success", "safetykoreachild"): os.getenv("R2_URL_DATAC_SUCCESS_10_SAFETYKOREACHILD"),
        ("dataC", "success", "safetykoreahome"): os.getenv("R2_URL_DATAC_SUCCESS_13_SAFETYKOREAHOME"),
        ("dataC", "success", "rra-cert"): os.getenv("R2_URL_DATAC_SUCCESS_11_RRA_CERT"),
        ("dataC", "success", "rra-self-cert"): os.getenv("R2_URL_DATAC_SUCCESS_12_RRA_SELF_CERT"),

        # DataC Failed 구조 매핑
        ("dataC", "failed", "safetykorea"): os.getenv("R2_URL_DATAC_FAILED_1_SAFETYKOREA"),
        ("dataC", "failed", "wadiz-makers"): os.getenv("R2_URL_DATAC_FAILED_2_WADIZ"),
        ("dataC", "failed", "efficiency-rating"): os.getenv("R2_URL_DATAC_FAILED_3_EFFICIENCY"),
        ("dataC", "failed", "high-efficiency"): os.getenv("R2_URL_DATAC_FAILED_4_HIGH_EFFICIENCY"),
        ("dataC", "failed", "standby-power"): os.getenv("R2_URL_DATAC_FAILED_5_STANDBY_POWER"),
        ("dataC", "failed", "approval"): os.getenv("R2_URL_DATAC_FAILED_6_APPROVAL"),
        ("dataC", "failed", "declaration-details"): os.getenv("R2_URL_DATAC_FAILED_7_DECLARE"),
        ("dataC", "failed", "kwtc"): os.getenv("R2_URL_DATAC_FAILED_8_KWTC"),
        ("dataC", "failed", "recall"): os.getenv("R2_URL_DATAC_FAILED_9_RECALL"),
        ("dataC", "failed", "safetykoreachild"): os.getenv("R2_URL_DATAC_FAILED_10_SAFETYKOREACHILD"),
        ("dataC", "failed", "safetykoreahome"): os.getenv("R2_URL_DATAC_FAILED_13_SAFETYKOREAHOME"),
        ("dataC", "failed", "rra-cert"): os.getenv("R2_URL_DATAC_FAILED_11_RRA_CERT"),
        ("dataC", "failed", "rra-self-cert"): os.getenv("R2_URL_DATAC_FAILED_12_RRA_SELF_CERT"),
    }

    r2_url = r2_url_mapping.get((category, subcategory))

    # 로컬 개발 환경 fallback (VERCEL 환경변수 없을 때)
    if not r2_url and os.getenv("VERCEL") is None:
        local_parquet_path = local_file_mapping.get(subcategory, "./parquet/1_safetykorea_flattened.parquet")
        if os.path.exists(local_parquet_path):
            logger.info(f"로컬 개발 환경: {category}/{subcategory} → {local_parquet_path}")
            return local_parquet_path
        else:
            logger.info(f"로컬 개발 환경: 기본 파일 사용")
            return "./parquet/1_safetykorea_flattened.parquet"

    if not r2_url:
        fallback_candidates = []
        local_candidate = local_file_mapping.get(subcategory)
        if local_candidate:
            fallback_candidates.extend([
                local_candidate,
                os.path.abspath(local_candidate),
                f"Project/{local_candidate}"
            ])

        for candidate in fallback_candidates:
            if candidate and os.path.exists(candidate):
                logger.warning(
                    f"R2 URL 누락: {category}/{subcategory} - 로컬 파일로 대체 ({candidate})"
                )
                return candidate

        raise ValueError(f"R2 URL not found for {category}/{subcategory}. Check environment variables.")

    logger.info(f"전체 데이터 모드: {category}/{subcategory} → R2")
    return r2_url

def search_in_fields(item: Dict[str, Any], keyword: str, search_field: str = "product_name") -> bool:
    """3개 검색 필드 지원 - '전체' 검색 기능 완전 제거"""
    keyword_lower = keyword.lower()
    
    # 다층 구조 처리 (SafetyKorea, Wadiz 등)
    if "resultData" in item:
        item = item["resultData"]
    elif "data" in item and isinstance(item["data"], dict):
        item = item["data"]
    
    # 검색 필드별 필드 변형들 (하드코딩)
    field_mappings = {
        "company_name": ["entrprsNm", "makerName", "importerName", "company_name", "제조자", "상호", "업체명", "maker_name", "사업자명"],
        "model_name": ["modelName", "model_name", "모델명"],
        "product_name": ["prductNm", "productName", "equipment_name", "제품명", "기자재명칭", "product_name", "품목명"]
    }
    
    if search_field in field_mappings:
        # 해당 필드의 모든 변형들에서 검색
        search_fields = field_mappings[search_field]
        
        for field in search_fields:
            if search_field_value(item, field, keyword_lower):
                return True
    
    return False

def search_nested_fields(item: Dict[str, Any], keyword: str) -> bool:
    """중첩된 구조에서 재귀적으로 검색"""
    for key, value in item.items():
        if isinstance(value, str):
            if keyword in value.lower():
                return True
        elif isinstance(value, list):
            for list_item in value:
                if isinstance(list_item, str) and keyword in list_item.lower():
                    return True
                elif isinstance(list_item, dict):
                    if search_nested_fields(list_item, keyword):
                        return True
        elif isinstance(value, dict):
            if search_nested_fields(value, keyword):
                return True
        elif value is not None:
            if keyword in str(value).lower():
                return True
    return False

def search_field_value(item: Dict[str, Any], field: str, keyword: str) -> bool:
    """특정 필드에서 키워드 검색"""
    if field in item:
        value = item[field]
        if isinstance(value, str) and keyword in value.lower():
            return True
        elif isinstance(value, list):
            for list_item in value:
                if isinstance(list_item, str) and keyword in list_item.lower():
                    return True
        elif value is not None and keyword in str(value).lower():
            return True
    return False

def generate_dynamic_summary(filtered_data: List[Dict], metadata: Dict) -> Dict[str, Any]:
    """동적 요약 정보 생성"""
    if not filtered_data:
        return {}
    
    summary = {
        "total_records": len(filtered_data),
        "data_source": metadata.get("source", "unknown"),
        "last_updated": metadata.get("last_updated", "unknown")
    }
    
    # 첫 번째 레코드의 필드로 분석 가능한 항목들 찾기
    if filtered_data:
        sample_item = filtered_data[0]
        
        # 날짜 필드 찾기 (date, 날짜, _date, _at 등이 포함된 필드)
        date_fields = [key for key in sample_item.keys() 
                      if any(date_keyword in key.lower() 
                            for date_keyword in ['date', '날짜', '_at', 'time'])]
        
        if date_fields:
            date_field = date_fields[0]  # 첫 번째 날짜 필드 사용
            dates = [item.get(date_field, "") for item in filtered_data if item.get(date_field)]
            if dates:
                summary["date_range"] = {
                    "field": date_field,
                    "earliest": min(dates),
                    "latest": max(dates)
                }
        
        # 카테고리성 필드 찾기 (category, type, status, 구분 등)
        category_fields = [key for key in sample_item.keys() 
                          if any(cat_keyword in key.lower() 
                                for cat_keyword in ['category', 'type', 'status', '구분', 'kind', 'class'])]
        
        for field in category_fields[:2]:  # 최대 2개 카테고리 필드만
            field_distribution = {}
            for item in filtered_data:
                value = item.get(field, "기타")
                field_distribution[value] = field_distribution.get(value, 0) + 1
            
            if field_distribution:
                summary[f"{field}_distribution"] = field_distribution
    
    return summary

async def filter_data_by_conditions(conditions: Dict[str, Any]) -> List[Dict[str, Any]]:
    """검색 조건으로 데이터 필터링 - 카테고리별 지원 (대용량 파일 지원)"""
    category = conditions.get("category")
    subcategory = conditions.get("subcategory")
    file_size_mb = 0  # 기본값 설정
    
    if category and subcategory:
        # 카테고리별 데이터 로드
        data_file_path = get_data_file_path(category, subcategory)
        if data_file_path:
            data_file_str, is_r2_url, is_tabular, file_size_mb = _inspect_data_source(data_file_path)

            if is_tabular or file_size_mb > 50:
                effective_subcategory = normalize_subcategory(subcategory)
                limit = conditions.get("limit", 1000000)
                offset = conditions.get("offset", 0)
                page = (offset // limit) + 1 if limit and limit > 0 else 1

                search_result = await duckdb_search_large_file(
                    file_path=data_file_str,
                    keyword=conditions.get("keyword"),
                    search_field=conditions.get("search_field", "all"),
                    limit=limit,
                    page=page,
                    filters=conditions.get("filters"),
                    category=category,
                    subcategory=effective_subcategory
                )

                if search_result.get("error"):
                    logger.warning(f"DuckDB 다운로드 실패, ijson으로 fallback: {search_result.get('message')}")
                    search_result = await stream_search_large_file(
                        file_path=data_file_path,
                        keyword=conditions.get("keyword"),
                        search_field=conditions.get("search_field", "all"),
                        filters=conditions.get("filters"),
                        limit=conditions.get("limit", None),
                        offset=conditions.get("offset", 0)
                    )

                filtered_data = search_result.get("results", [])
            else:
                with open(data_file_path, 'r', encoding='utf-8') as f:
                    category_data = json.load(f)
                filtered_data = category_data.get("data", [])

                if conditions.get("keyword"):
                    keyword = conditions["keyword"].lower()
                    filtered_data = [
                        item for item in filtered_data
                        if search_in_all_fields(item, keyword)
                    ]
        else:
            filtered_data = []
    else:
        # 기본 더미 데이터 로드
        dummy_data = load_dummy_data()
        filtered_data = dummy_data["data"]
        
        # 키워드 검색
        if conditions.get("keyword"):
            keyword = conditions["keyword"].lower()
            filtered_data = [
                item for item in filtered_data
                if keyword in item.get("title", "").lower() or
                   keyword in item.get("content", "").lower() or
                   any(keyword in tag.lower() for tag in item.get("tags", []))
            ]
    
    # 추가 필터들 (대용량 파일은 스트리밍에서 이미 처리됨)
    if not (category and subcategory and file_size_mb > 50):
        if conditions.get("categories"):
            filtered_data = [
                item for item in filtered_data
                if item.get("category") in conditions["categories"]
            ]
        
        # 날짜 필터
        if conditions.get("date_from"):
            filtered_data = [
                item for item in filtered_data
                if item.get("date", "") >= conditions["date_from"]
            ]
        
        if conditions.get("date_to"):
            filtered_data = [
                item for item in filtered_data
                if item.get("date", "") <= conditions["date_to"]
            ]
    
    return filtered_data

def search_in_all_fields(item: Dict[str, Any], keyword: str) -> bool:
    """모든 필드에서 키워드 검색"""
    # resultData 구조 처리
    if "resultData" in item:
        item = item["resultData"]
    
    for key, value in item.items():
        if isinstance(value, str):
            if keyword in value.lower():
                return True
        elif isinstance(value, list):
            for list_item in value:
                if isinstance(list_item, str) and keyword in list_item.lower():
                    return True
        elif value is not None:
            if keyword in str(value).lower():
                return True
    return False

def infer_field_type(sample_data: List[Dict], field_name: str) -> str:
    """샘플 데이터에서 필드 타입 추론"""
    if not sample_data:
        return "text"
    
    sample_values = []
    for item in sample_data:
        # resultData 구조 처리
        if "resultData" in item:
            item = item["resultData"]
        
        if field_name in item and item[field_name] is not None:
            sample_values.append(item[field_name])
    
    if not sample_values:
        return "text"
    
    # 날짜 형식 검사
    if any(keyword in field_name.lower() for keyword in ['date', '날짜', 'time', '_at']):
        return "date"
    
    # 숫자 형식 검사
    if all(isinstance(val, (int, float)) for val in sample_values if val is not None):
        return "number"
    
    # 배열 형식 검사
    if any(isinstance(val, list) for val in sample_values):
        return "array"
    
    # URL 검사
    if any(isinstance(val, str) and (val.startswith('http') or val.startswith('www')) for val in sample_values):
        return "link"
    
    # 이미지 파일 확장자 검사
    if any(isinstance(val, str) and any(val.lower().endswith(ext) for ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']) for val in sample_values):
        return "image"
    
    return "text"

def get_display_config_for_rendering(category: str, subcategory: str) -> Dict[str, Any]:
    """렌더링용 표시 설정 조회"""
    try:
        config = display_config_manager.export_client_config(category, subcategory)
        return config
    except:
        return {}

# 설정 시스템 초기화
def initialize_field_settings():
    """애플리케이션 시작시 field_settings.json 초기화"""
    try:
        # field_settings.json 파일이 비어있거나 없으면 기본 설정 생성
        field_settings_path = Path(__file__).parent.parent / "config" / "field_settings.json"
        
        if not field_settings_path.exists() or field_settings_path.stat().st_size < 100:
            logger.info("기본 field_settings.json 파일을 생성합니다.")
            # 파일이 이미 생성되어 있으므로 추가 작업 없음
        
        # display_config_manager 새로고침
        display_config_manager._field_settings = display_config_manager._load_field_settings()
        display_config_manager._configs.clear()
        display_config_manager._load_all_configs()
        
        logger.info("설정 시스템 초기화 완료")
        
    except Exception as e:
        logger.error(f"설정 시스템 초기화 실패: {e}")



# 애플리케이션 시작시 초기화
initialize_field_settings()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
# 프리페치 구성을 조회하는 헬퍼 (확장 가능)
def get_prefetch_config() -> Dict[str, Any]:
    data_mode = os.getenv("DATA_MODE", "full").lower()
    enabled = data_mode == "2025"
    return {
        "enabled": enabled,
        "data_mode": data_mode,
        "supported_keys": list(BLOB_ENV_PREFETCH_MAPPING.keys())
    }
